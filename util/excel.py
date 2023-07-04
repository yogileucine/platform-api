import openpyxl


class Excel:
    """ This method is used to read excel by column name and row number, and it returns cell value """

    @staticmethod
    def read_excel_col_name_row_number(filename, sheet_name, column_name, row_number):
        wb1 = openpyxl.load_workbook(filename, data_only=True)
        sheet = wb1[sheet_name]
        col = 1
        for i in range(1, sheet.max_column + 1):
            cell_obj = sheet.cell(row=1, column=i)
            if cell_obj.value == column_name:
                col = i
                break
        cell_obj = sheet.cell(row=row_number, column=col)
        return cell_obj.value

    """ This method is used to read excel by column name and row value 
    and it takes input parameter as filename,sheet name, list of columns, row value 
    and it returns data in dictionary format like key and value as list example {"data": ["one", "two"]}"""

    @staticmethod
    def read_excel_columns_name_row_value(filename, sheet_name, columns, row_value):
        wb = openpyxl.load_workbook(filename, data_only=True)
        sheet = wb[sheet_name]
        data = {}
        col_num = []
        for i in columns:
            col_found = False
            for j in range(1, sheet.max_column + 1):
                cell_obj = sheet.cell(row=1, column=j)
                if cell_obj.value == i:
                    col_num.append(j)
                    col_found = True
                    break
            if col_found is False:
                col_num.append(0)
        for i in range(0, len(columns)):
            r = 2
            if col_num[i] == 0:
                data.update({columns[i]: None})
                continue
            cell_obj1 = sheet.cell(row=r, column=col_num[i])
            cell_data = []
            while cell_obj1.value is not None:
                if sheet.cell(row=r, column=1).value == row_value:
                    cell_data.append(cell_obj1.value)
                r = r + 1
                cell_obj1 = sheet.cell(row=r, column=col_num[i])
            data.update({columns[i]: cell_data})
        return data

    """ This method is used to read excel by columns name 
        and it takes input parameter as list of columns
        and it returns data in dictionary format like key and value as list example {"data": ["one", "two"]} """

    @staticmethod
    def read_excel_columns(filename, sheet_name, columns):
        wb1 = openpyxl.load_workbook(filename, data_only=True)
        sheet = wb1[sheet_name]
        data = {}
        col_num = []
        for i in columns:
            col_found = False
            for j in range(1, sheet.max_column + 1):
                cell_obj = sheet.cell(row=1, column=j)
                if cell_obj.value == i:
                    col_num.append(j)
                    col_found = True
                    break
            if col_found is False:
                col_num.append(0)
        for i in range(0, len(columns)):
            r = 2
            if col_num[i] == 0:
                data.update({columns[i]: None})
                continue
            cell_obj1 = sheet.cell(row=r, column=col_num[i])
            cell_data = []
            while cell_obj1.value is not None:
                cell_data.append(cell_obj1.value)
                r = r + 1
                cell_obj1 = sheet.cell(row=r, column=col_num[i])
            data.update({columns[i]: cell_data})
        return data
